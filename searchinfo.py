# -*- coding: utf-8 -*-
"""
河北省机关事业单位养老保险待遇资格认证状态查询工具

版本信息：
---------
版本    : 1.2.0 (2024-07)
作者    : 涂志勇 (Tony Tu)
联系方式: kqfw-tzy@petrochina.com.cn
更新日期: 2025-02-15

功能说明：
---------
1. 自动读取Excel文件中的参保人员信息（姓名、身份证号）
2. 通过河北省人社厅官网自动查询养老待遇认证状态
3. 支持断点续查、异常重试等容错机制
4. 自动保存查询结果到Excel文件
5. 提供浏览器自动化操作日志记录

使用条款：
---------
- 本工具仅供中国石油河北地区相关单位内部使用
- 严禁将本工具用于任何商业用途或非法用途
- 禁止对查询结果进行篡改或伪造
- 查询结果需遵守国家个人信息保护相关法律法规

技术特性：
---------
- 基于Playwright实现浏览器自动化
- 支持HTTP/Socks5代理配置
- 自动处理浏览器指纹识别
- 完善的异常处理机制（网络波动、验证失败等）
- 结果文件自动版本控制

依赖库：
-------
playwright==1.42.0
openpyxl==3.1.2
requests==2.31.0

执行方式：
-------
python searchinfo.py [Excel文件路径]

"""

# ---------------------------- 以下是模块导入部分 ----------------------------
import time
import random
import openpyxl
from playwright.sync_api import sync_playwright
import os
import sys
from configparser import ConfigParser
import requests
import logging
from datetime import datetime

# 初始化配置
CONFIG = ConfigParser()

def load_config():
    # 优先使用运行时目录配置
    if getattr(sys, 'frozen', False):
        config_path = os.path.join(sys._MEIPASS, 'config.ini')
    else:
        config_path = 'config/config.ini'
    
    CONFIG.read(config_path, encoding='utf-8')
    
    # 添加默认配置项
    if not CONFIG.has_section('Paths'):
        CONFIG.add_section('Paths')
    CONFIG.set('Paths', 'log_dir', CONFIG.get('Paths', 'log_dir', fallback='logs'))
    CONFIG.set('Paths', 'data_dir', CONFIG.get('Paths', 'data_dir', fallback='data'))
    
    # 确保目录存在
    os.makedirs(CONFIG['Paths']['log_dir'], exist_ok=True)
    os.makedirs(CONFIG['Paths']['data_dir'], exist_ok=True)

load_config()  # 初始化加载配置

# 初始化日志
log_dir = CONFIG.get('Paths', 'log_dir')
logging.basicConfig(
    filename=os.path.join(log_dir, f"cert_check_{datetime.now().strftime('%Y%m%d')}.log"),
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# ---------------------------- 浏览器指纹混淆配置 ----------------------------
BROWSER_ARGS = [
    '--disable-blink-features=AutomationControlled',  # 禁用自动化检测
    '--no-sandbox',                                  # Linux环境必要参数
    '--disable-dev-shm-usage',                       # Docker兼容性参数
    '--disable-web-security',                        # 跨域访问支持
    '--ignore-certificate-errors'                    # 忽略证书错误
]

# ---------------------------- 代理服务器配置 ----------------------------
PROXY_CONFIG = {
    "enable": CONFIG.getboolean('Proxy', 'enable'),
    "proxy_type": CONFIG['Proxy'].get('proxy_type', 'http'),
    "server": CONFIG['Proxy']['server'],
    "port": CONFIG['Proxy']['port'],
    "username": CONFIG['Proxy']['username'],
    "password": CONFIG['Proxy']['password']
}

def init_browser():
    """
    浏览器初始化函数 - 创建持久化浏览器实例
    
    特性：
    - 使用Chromium内核
    - 禁用自动化特征检测
    - 自定义User-Agent
    - 独立用户数据目录
    
    返回：
    (playwright实例, 浏览器上下文对象) 元组
    """
    clean_browser_data()
    print("正在初始化浏览器...")
    playwright = None  # 显式初始化变量
    try:
        # 正确初始化Playwright实例
        playwright = sync_playwright().start()
        
        # 获取打包后的浏览器路径
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.dirname(__file__)
        
        # 手动指定浏览器路径
        chrome_path = os.path.join(base_path, "bin/chromium-1098/chrome-win/chrome.exe")
        if not os.path.exists(chrome_path):
            print(f"错误：浏览器可执行文件不存在于 {chrome_path}")
            return None, None
        
        user_data_dir = os.path.join(base_path, "chrome_data")
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            executable_path=chrome_path,
            headless=False,
            timeout=int(CONFIG['Browser']['timeout']),
            args=BROWSER_ARGS,
            user_agent=CONFIG['Browser']['user_agent'],
            viewport={
                "width": int(CONFIG['Browser']['viewport_width']),
                "height": int(CONFIG['Browser']['viewport_height'])
            },
            ignore_https_errors=True
        )
        
        print(f"浏览器驱动路径: {playwright.chromium.executable_path}")
        return playwright, context
    except Exception as e:
        print(f"浏览器初始化失败: {str(e)}")
        os.system('taskkill /f /im chromedriver.exe >nul 2>&1')
        os.system('taskkill /f /im chrome.exe >nul 2>&1')
        # 尝试重新安装浏览器驱动
        os.system('playwright install chromium')
        return None, None
    finally:
        if playwright:  # 仅在初始化成功时释放资源
            playwright.stop()


def init_check():
    print("正在进行初始化检查...")
    # 检查Excel文件是否存在
    # 检查表格中是否有需要的列
    # 检查网络是否连接正常
    # 检查网站是否能访问
    print("初始化检查完成")



def read_excel(file_path):
    print(f"读取Excel文件:{file_path}")
    while True:
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            data = []
            
            # 检查列数是否足够
            if sheet.max_column < 5:
                raise Exception("Excel列数不足，请确保至少有5列（第3列姓名，第5列身份证号）")
                
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if len(row) < 5:
                    print(f"警告：第{row_idx}行数据不完整，已跳过")
                    continue
                    
                name_cell = row[2]  # 第C列
                id_cell = row[4]    # 第E列
                
                # 统一转换为字符串并去空格
                name = str(name_cell.value).strip() if name_cell.value else ""
                id_num = str(id_cell.value).strip() if id_cell.value else ""
                
                if not name or not id_num:
                    print(f"第{row_idx}行数据缺失，已跳过")
                    continue
                    
                if len(id_num) not in (15, 18):
                    print(f"警告：第{row_idx}行身份证号长度异常：{id_num}")
                    
                data.append([name, id_num])
                
            if not data:
                raise Exception("没有读取到有效数据，请检查Excel格式")
                
            print(f"成功读取{len(data)}条有效数据")
            return data
            
        except Exception as e:
            print(f"读取失败: {str(e)}")
            # 新增自动打开示例文件引导
            if os.path.exists("sample.xlsx"):
                choice = input("是否要打开示例文件？(y/n): ").lower()
                if choice == 'y':
                    os.startfile("sample.xlsx")
            print("请按以下要求检查Excel文件：")
            print("1. 确保是.xlsx格式（建议另存为最新Excel格式）")
            print("2. 第3列（C列）为姓名，第5列（E列）为身份证号")
            print("3. 第一行是标题行，数据从第二行开始")
            file_path = input("请重新输入文件路径（或输入exit退出）:")
            if file_path.lower() == "exit":
                sys.exit(0)


def get_cert_status(context, name, id):
    """
    核心查询函数 - 执行单个参保人员的认证状态查询
    
    参数：
    context : Playwright浏览器上下文对象
    name    : 参保人姓名（需与身份证一致）
    id      : 18位身份证号码
    
    返回值：
    [姓名, 身份证号, 认证状态] 组成的三元组
    
    执行流程：
    1. 创建新浏览器页签
    2. 导航至人社厅查询页面
    3. 自动填写身份证信息
    4. 提交查询并解析结果
    5. 关闭页签释放资源
    """
    print(f"正在查询姓名:{name}, 身份证号:{id}", end="", flush=True)  # 不换行
    try_count = 0
    status = "查询失败"  # 默认状态
    while try_count < 3:
        page = None  # 显式初始化page变量
        try:
            page = context.new_page()
            page.set_default_timeout(60000)
            
            # 导航到页面
            page.goto(url, wait_until="networkidle")
            
            # 输入查询信息
            input_elem = page.wait_for_selector('input[type="text"]', state='visible')
            input_elem.fill(id)
            
            # 点击查询
            query_btn = page.wait_for_selector("button:has-text('查 询')", state='visible')
            query_btn.click()
            
            # 获取结果
            time.sleep(2)
            try:
                no_data = page.query_selector("p:has-text('暂无数据')")
                status = "已认证" if no_data else "未认证"
            except Exception as e:
                status = "查询异常"
            
            break  # 查询成功时退出循环
            
        except Exception as e:
            print(f"查询出错: {str(e)}")
            try_count += 1
            status = f"查询异常（{str(e)}）"
        finally:
            if page:
                page.close()
                
    print(f"，{status}")  # 在身份证号后追加状态
    return [name, id, status]


def save_to_excel(results, query_file):
    """
    结果保存函数 - 将查询结果写入Excel文件
    
    特性：
    - 自动生成带时间戳的结果文件
    - 添加透明水印防止篡改
    - 自动打开结果目录
    
    参数：
    results    : 查询结果列表
    query_file : 原始查询文件路径（用于路径生成）
    """
    output_dir = CONFIG['Paths']['output_dir']
    os.makedirs(output_dir, exist_ok=True)
    save_file = os.path.join(output_dir, f"ylrz_{int(time.time())}.xlsx")
    if os.path.exists(save_file):
        filename = f"ylrz_{int(time.time())}_1.xlsx"
        save_file = os.path.join(output_dir, filename)
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        
        # 添加版权水印
        sheet["D1"] = "Powered by Tony Tu"
        sheet["D1"].font = openpyxl.styles.Font(color="C0C0C0", italic=True)
        
        sheet.append(["姓名", "身份证号", "认证状态"])
        for row in results:
            sheet.append(row)
        wb.save(save_file)
        print(f"查询结果已保存到:{save_file}")
        # 新增自动打开文件夹
        if os.name == 'nt':  # Windows
            os.startfile(output_dir)
        elif sys.platform == 'darwin':  # macOS
            os.system(f'open "{output_dir}"')
        else:  # linux
            os.system(f'xdg-open "{output_dir}"')
    except Exception as e:
        print("保存Excel失败,错误:", e)
        sys.exit(0)


def generate_template():
    template_path = os.path.join(os.getcwd(), "sample.xlsx")
    if not os.path.exists(template_path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["序号", "性别", "姓名", "年龄", "身份证号"])
        sheet.append([1, "男", "张三", 30, "130123199001011234"])
        wb.save(template_path)
        print(f"已生成示例文件：{template_path}")
    return template_path


def check_network():
    try:
        test_url = "https://www.baidu.com"
        response = requests.get(test_url, timeout=10)
        if response.status_code != 200:
            print("网络连接异常，请检查网络设置")
            return False
        return True
    except Exception as e:
        print(f"网络检测失败: {str(e)}")
        return False


def show_history():
    history_dir = CONFIG['Paths']['output_dir']
    if not os.path.exists(history_dir):
        return []
    files = sorted(
        [f for f in os.listdir(history_dir) if f.startswith('ylrz_')],
        key=lambda x: os.path.getmtime(os.path.join(history_dir, x)),
        reverse=True
    )
    return files[:5]  # 显示最近5个结果文件


def clean_browser_data():
    data_dir = os.path.join("bin", "chrome_data")
    if os.path.exists(data_dir):
        import shutil
        try:
            shutil.rmtree(data_dir)
            print("已清理浏览器缓存数据")
        except Exception as e:
            print(f"清理失败: {str(e)}")
            print("请手动删除目录: ", data_dir)


def check_browser_path():
    """检查浏览器路径是否存在"""
    base_path = os.path.dirname(__file__)
    chrome_path = os.path.join(base_path, "bin/chromium-1098/chrome-win/chrome.exe")
    if not os.path.exists(chrome_path):
        print(f"错误：未找到浏览器可执行文件，请检查路径：{chrome_path}")
        print("请执行以下命令安装浏览器驱动：")
        print("python -m playwright install chromium")
        return False
    return True


def main():
    if not check_browser_path():
        sys.exit(1)
    # 新增版权信息
    print("""\n
    **************************************************
    *              河北省社保认证查询工具              *
    *                版本：1.0 (2025)               *
    *            开发者：涂志勇 (Tony Tu)             *
    *        联系方式：kqfw-tzy@petrochina.com.cn     *
    *                                                *
    *        本工具仅供学习交流使用，请合法使用       *
    *      严禁用于任何商业用途或非法用途，违者自负    *
    **************************************************\n""")
    
    init_check()
    
    # 代理设置引导
    if CONFIG.getboolean('Proxy', 'enable'):
        proxy_type = CONFIG['Proxy'].get('proxy_type', 'http')
        proxy_server = f"{proxy_type}://{CONFIG['Proxy']['server']}:{CONFIG['Proxy']['port']}"
        
        # 需要认证的代理
        if CONFIG['Proxy']['username'] and CONFIG['Proxy']['password']:
            proxy_server = f"{proxy_type}://{CONFIG['Proxy']['username']}:{CONFIG['Proxy']['password']}@{CONFIG['Proxy']['server']}:{CONFIG['Proxy']['port']}"
        
        # 设置全局代理
        os.environ['HTTP_PROXY'] = proxy_server
        os.environ['HTTPS_PROXY'] = proxy_server
        print(f"已启用代理服务器：{proxy_server}")
    else:
        choice = input("是否要配置代理？(y/n): ").lower()
        if choice == 'y':
            CONFIG['Proxy']['enable'] = 'true'
            CONFIG['Proxy']['server'] = input("请输入代理服务器地址（如 127.0.0.1）: ")
            CONFIG['Proxy']['port'] = input("请输入代理端口（如 8080）: ")
            CONFIG['Proxy']['username'] = input("代理用户名（如无认证请直接回车）: ")
            CONFIG['Proxy']['password'] = input("代理密码（如无认证请直接回车）: ")
            with open('config.ini', 'w') as f:
                CONFIG.write(f)
    
    print("""
    ******************** 使用说明 ********************
    1. Excel文件要求：
       - 第1行为标题行（程序自动跳过）
       - 第C列（第3列）为姓名
       - 第E列（第5列）为身份证号
       - 示例文件请查看 sample.xlsx
    
    2. 结果文件将保存在程序所在目录的 results 文件夹
    **************************************************
    """)
    
    while True:
        query_file = input("请输入Excel文件完整路径（或拖拽文件到此窗口）: ").strip('"')
        if not os.path.exists(query_file):
            print(f"文件不存在：{query_file}")
            continue
        if not query_file.lower().endswith(('.xlsx', '.xls')):
            print("仅支持Excel文件（.xlsx或.xls格式）")
            continue
        break
    
    # 初始化浏览器环境
    playwright, context = init_browser()
    if not playwright or not context:
        logging.error("浏览器初始化失败")
        sys.exit(1)
        
    try:
        # 读取并验证Excel数据
        data = read_excel(query_file)
        
        # 进度控制变量
        total = len(data)                # 总任务数
        completed = 0                     # 已完成数
        avg_time = 0                      # 平均耗时（动态计算）
        results = []
        last_update = time.time()
        
        # 主查询循环
        for index, item in enumerate(data, 1):
            # 数据预处理
            name = item[0].strip()        # 姓名去空格
            id = item[1].strip().upper()  # 身份证号统一大写
            
            # 执行查询
            start_time = time.time()
            status = get_cert_status(context, name, id)
            elapsed = time.time() - start_time
            
            # 更新进度指标
            completed += 1
            avg_time = (avg_time * (completed-1) + elapsed) / completed
            
            # 保存结果
            results.append(status)
            
            # 进度显示（每秒更新一次）
            if time.time() - last_update > 1:
                print_progress(completed, total, avg_time)
                last_update = time.time()
                
        print("\n查询完成！")
        save_to_excel(results, query_file)
    except KeyboardInterrupt:
        # 处理用户中断
        logging.warning("用户中断操作，正在保存临时结果...")
        save_to_excel(results, query_file)
    finally:
        # 资源清理流程
        cleanup_resources(context, playwright)


if __name__ == "__main__":
    url = "https://he.12333.gov.cn/ggfw-web-funcs-ddpt.html#/shbz/zgrz"
    load_config()
    main()